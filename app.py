import base64
import json
import os
import re
import sqlite3
import urllib.parse
from datetime import datetime
from io import BytesIO

import pandas as pd
import streamlit as st
from PIL import Image

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Fallback structures for external parsing libraries
try:
    import fitz
except Exception:
    fitz = None

try:
    import pypdfium2 as pdfium
except Exception:
    pdfium = None

try:
    from pdf2image import convert_from_bytes
except Exception:
    convert_from_bytes = None

try:
    import google.generativeai as genai
except Exception:
    genai = None

try:
    from openai import OpenAI
except Exception:
    OpenAI = None

try:
    from google.cloud import vision
except Exception:
    vision = None

import warnings
warnings.filterwarnings("ignore")

# Global System Constants
APP_TITLE = "Deep CSC - AI Multi-Bill Processor Ultra"
DB_PATH = "bills.db"
OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

def secret_or_default(key, default=""):
    try:
        return st.secrets[key]
    except Exception:
        return default

DEFAULT_USERNAME = secret_or_default("APP_USERNAME", "admin")
DEFAULT_PASSWORD = secret_or_default("APP_PASSWORD", "password123")

def setup_page():
    st.set_page_config(page_title=APP_TITLE, page_icon="🧾", layout="wide", initial_sidebar_state="expanded")

def init_db():
    with sqlite3.connect(DB_PATH, timeout=30) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS bills (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                shop_name TEXT NOT NULL,
                bill_date TEXT NOT NULL,
                gst_number TEXT,
                total REAL NOT NULL,
                calculated_total REAL NOT NULL,
                status TEXT NOT NULL,
                timestamp TEXT NOT NULL,
                UNIQUE(shop_name, bill_date, total) ON CONFLICT IGNORE
            )
            """
        )

def init_auth():
    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False

def init_runtime_state():
    defaults = {
        "selected_provider": "Gemini",
        "theme_mode": "light",
        "processed_files": set(),
        "gemini_available": True,
        "last_gemini_error_time": None,
        "gemini_cooldown_seconds": 900,
        "current_results": []
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

def do_login():
    st.markdown("""
        <style>
        .login-wrap {
            max-width: 460px;
            margin: 7vh auto;
            padding: 28px;
            border-radius: 24px;
            background: rgba(255,255,255,0.85);
            box-shadow: 0 20px 60px rgba(15, 23, 42, 0.14);
            border: 1px solid rgba(148,163,184,0.18);
            backdrop-filter: blur(10px);
        }
        .login-title {
            font-size: 32px;
            font-weight: 800;
            margin: 0;
            background: linear-gradient(to right, #0f172a, #4338ca);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
        }
        .login-sub { margin-top: 8px; color: #64748b; font-size: 14px; }
        </style>
    """, unsafe_allow_html=True)

    st.markdown('<div class="login-wrap">', unsafe_allow_html=True)
    st.markdown('<div class="login-title">🔐 System Login</div>', unsafe_allow_html=True)
    st.markdown('<div class="login-sub">Secure access for AI bill dashboard</div>', unsafe_allow_html=True)

    username = st.text_input("Username")
    password = st.text_input("Password", type="password")
    if st.button("Login Server", use_container_width=True, key="login_btn"):
        if username == DEFAULT_USERNAME and password == DEFAULT_PASSWORD:
            st.session_state.logged_in = True
            st.rerun()
        else:
            st.error("Invalid Username or Password Credentials")
    st.markdown("</div>", unsafe_allow_html=True)
    st.stop()

def terminate_session():
    st.session_state.logged_in = False
    for k in list(st.session_state.keys()):
        if k != "logged_in":
            del st.session_state[k]
    st.rerun()

def apply_theme_css():
    theme_mode = st.session_state.get("theme_mode", "light")
    if theme_mode == "dark":
        st.markdown("""
            <style>
            .stApp { background: #0b1220; color: #e5e7eb; }
            section[data-testid="stSidebar"] { background: #0f172a; }
            section[data-testid="stSidebar"] * { color: #f8fafc !important; }
            .stButton>button { background: linear-gradient(135deg, #4f46e5 0%, #2563eb 100%) !important; color: white !important; }
            </style>
        """, unsafe_allow_html=True)
    else:
        st.markdown("""
            <style>
            .stApp { background: #f8fafc; color: #0f172a; }
            section[data-testid="stSidebar"] { background: #0f172a; }
            section[data-testid="stSidebar"] * { color: #f8fafc !important; }
            </style>
        """, unsafe_allow_html=True)

def apply_css():
    st.markdown("""
        <style>
        .stApp { background: radial-gradient(circle at top left, #ffffff 0%, #eef2ff 45%, #e2e8f0 100%); }
        .deep-csc-header {
            background: linear-gradient(135deg, #0f172a 0%, #1e1b4b 45%, #312e81 100%);
            padding: 28px 30px; border-radius: 28px; margin-bottom: 18px;
            border: 1px solid rgba(255,255,255,0.10); box-shadow: 0 18px 40px rgba(15, 23, 42, 0.22);
            display: flex; align-items: center; justify-content: space-between; gap: 18px; flex-wrap: wrap;
        }
        .branding-text h1 {
            margin: 0; font-size: 34px !important; font-weight: 800;
            background: linear-gradient(to right, #38bdf8, #c084fc, #f43f5e);
            -webkit-background-clip: text; -webkit-text-fill-color: transparent;
        }
        .branding-text p { margin: 6px 0 0 0; color: #cbd5e1; font-size: 14px; }
        .branding-badge {
            background: linear-gradient(135deg, #ec4899 0%, #8b5cf6 100%); color: white !important;
            padding: 9px 16px; border-radius: 999px; font-size: 12px !important; font-weight: 700; text-transform: uppercase;
        }
        .csc-meta-badge {
            background: rgba(255,255,255,0.08); border: 1px solid rgba(255,255,255,0.14);
            color: #e2e8f0 !important; padding: 12px 16px; border-radius: 16px; font-size: 13px !important;
        }
        .metric-card {
            background: rgba(255,255,255,0.82); border: 1px solid rgba(148,163,184,0.18);
            border-radius: 22px; padding: 18px 20px; box-shadow: 0 14px 35px rgba(15, 23, 42, 0.10);
        }
        .metric-label { color: #64748b; font-size: 13px; margin-bottom: 8px; }
        .metric-value { font-size: 28px; font-weight: 800; color: #0f172a; }
        .metric-sub { color: #64748b; font-size: 13px; }
        .section-card { background: rgba(255,255,255,0.78); border: 1px solid rgba(148,163,184,0.16); border-radius: 22px; padding: 18px; }
        .stButton > button {
            background: linear-gradient(135deg, #4f46e5 0%, #2563eb 100%) !important;
            color: white !important; font-weight: 700 !important; border-radius: 14px !important;
        }
        .stButton > button p { color: white !important; }
        .stTabs [data-baseweb="tab-list"] { gap: 10px; background: rgba(255,255,255,0.55); padding: 8px; border-radius: 18px; }
        .stTabs [data-baseweb="tab"] { border-radius: 14px; padding: 10px 18px; font-weight: 700; }
        .stTabs [aria-selected="true"] { background: linear-gradient(135deg, #0f172a 0%, #4338ca 100%) !important; color: white !important; }
        div[data-testid="stDataFrame"] { border-radius: 18px; overflow: hidden; }
        section[data-testid="stSidebar"] { background: linear-gradient(180deg, #0f172a 0%, #111827 100%); }
        section[data-testid="stSidebar"] * { color: #f8fafc !important; }
        .block-container { padding-top: 1.1rem; padding-bottom: 2rem; }
        </style>
    """, unsafe_allow_html=True)

def metric_card(label, value, sub=""):
    st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">{label}</div>
            <div class="metric-value">{value}</div>
            <div class="metric-sub">{sub}</div>
        </div>
    """, unsafe_allow_html=True)

def render_metrics(df):
    c1, c2, c3, c4 = st.columns(4)
    total_files = len(df) if df is not None and not df.empty else 0
    matched = int((df["status"] == "Matched").sum()) if total_files else 0
    mismatch = int((df["status"] == "Mismatch").sum()) if total_files else 0
    review = int((df["status"] == "Needs Review").sum()) if total_files else 0

    with c1: metric_card("Total Bills Found", total_files, "Extracted from batch")
    with c2: metric_card("Matched Bills", matched, "Sum verified perfectly")
    with c3: metric_card("Mismatch / Review", mismatch + review, "Requires verification")
    with c4: metric_card("OCR Provider", st.session_state.get("selected_provider", "Gemini"), "Active pipeline")

@st.cache_resource
def setup_gemini():
    if genai is None: return None
    api_key = secret_or_default("GEMINI_API_KEY", "").strip()
    if not api_key: return None
    genai.configure(api_key=api_key)
    return genai.GenerativeModel("gemini-2.5-flash")

@st.cache_resource
def setup_openai():
    api_key = secret_or_default("OPENAI_API_KEY", "").strip()
    if not api_key or OpenAI is None: return None
    return OpenAI(api_key=api_key)

@st.cache_resource
def setup_perplexity():
    api_key = secret_or_default("PERPLEXITY_API_KEY", "").strip()
    if not api_key or OpenAI is None: return None
    return OpenAI(api_key=api_key, base_url="https://api.perplexity.ai")

@st.cache_resource
def setup_google_vision():
    return vision.ImageAnnotatorClient() if vision is not None else None

def validate_gst(gst_str):
    if not gst_str: return False, "N/A"
    gst_regex = r"^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$"
    clean_gst = re.sub(r"[^A-Z0-9]", "", str(gst_str).upper())
    return bool(re.match(gst_regex, clean_gst)), clean_gst

def normalize_items(items):
    if not isinstance(items, list): return []
    cleaned = []
    for it in items:
        if not isinstance(it, dict): continue
        qty_str = str(it.get("qty", "")).strip()
        if qty_str in ["", "None", "null", "NULL", "false", "NaN"]:
            qty_str = "1"
        else:
            num_match = re.search(r"[-+]?\d*\.\d+|\d+", qty_str)
            if num_match: qty_str = num_match.group(0)
            else: qty_str = "1"

        rate = str(it.get("rate", "")).strip()
        amount = str(it.get("amount", "")).strip()

        if not amount and qty_str and rate:
            try: amount = str(float(qty_str) * float(rate))
            except Exception: amount = "0"

        cleaned.append({
            "name": it.get("name") or "Unknown Item",
            "qty": qty_str,
            "rate": rate or "0",
            "amount": amount or "0"
        })
    return cleaned

def parse_json_from_response(response_text):
    raw = (response_text or "").strip().replace("```json", "").replace("```", "").strip()
    m = re.search(r"\{.*\}|\[.*\]", raw, re.DOTALL)
    if m: raw = m.group(0)
    parsed = json.loads(raw)
    if isinstance(parsed, list):
        return {"bills": parsed}
    if isinstance(parsed, dict) and "bills" not in parsed:
        if "items" in parsed or "shop_name" in parsed:
            return {"bills": [parsed]}
    return parsed

def safe_float(value):
    if not value: return 0.0
    try: return float(str(value).replace("₹", "").replace(",", "").replace(" ", "").strip())
    except Exception: return 0.0

def can_try_gemini():
    if st.session_state.get("gemini_available", True): return True
    last_error = st.session_state.get("last_gemini_error_time")
    return (not last_error) or ((datetime.now() - last_error).total_seconds() >= st.session_state.get("gemini_cooldown_seconds", 900))

def is_gemini_quota_error(err):
    msg = str(err).lower()
    return ("429" in msg) or ("quota" in msg) or ("resource_exhausted" in msg) or ("rate limit" in msg)

def build_schema_prompt():
    return """
You are an expert multi-invoice intelligence extraction agent. 
An inherent capability to detect MULTIPLE separate bills or receipts from a single page image is expected.
Analyze the provided document image and look closely for horizontal/vertical demarcations, distinct header zones, or multiple standalone lists.

Return ONLY a valid JSON object matching the exact schema below. Do not include any explanation or backticks.

{
  "bills": [
    {
      "shop_name": string or null,
      "bill_date": string or null,
      "gst_number": string or null,
      "items": [{"name": string, "qty": string, "rate": string, "amount": string}],
      "total": string or null
    }
  ]
}

Extraction Strategy Rules:
1. "bills" is an array. If there is 1 bill, return 1 object inside the array. If there are multiple distinct bills captured on the page, split them cleanly into separate structural entities inside the array.
2. For each bill, look for line items. Ensure you extract the item name, accurate quantity, rate, and total amount.
3. CRITICAL QUANTITY RULE: Quantity must NEVER be blank or non-numeric. Look closely at columns like "Qty", "Pcs", "Nos", "Pkt", "Quantity", "मात्रा". If a specific quantity is omitted or represented implicitly, you MUST intelligently infer it or use "1" as a structural fallback. Never leave it empty.
4. Extract every single line item visible on the document. Do not truncate.
"""

def image_to_bytes(image):
    buf = BytesIO()
    image.save(buf, format="JPEG", quality=95)
    return buf.getvalue()

def preprocess_for_ocr(image):
    return image

def convert_pdf_to_images(file_bytes):
    if fitz is not None:
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        images = []
        for page in doc:
            pix = page.get_pixmap(matrix=fitz.Matrix(4, 4), alpha=False)
            images.append(Image.frombytes("RGB", [pix.width, pix.height], pix.samples))
        doc.close()
        return images
    if pdfium is not None:
        pdf = pdfium.PdfDocument(file_bytes)
        images = []
        for i in range(len(pdf)):
            images.append(pdf[i].render(scale=4).to_pil().convert("RGB"))
        return images
    if convert_from_bytes is not None:
        return convert_from_bytes(file_bytes, dpi=400)
    raise RuntimeError("No PDF rendering library available.")

def extract_vision_text(vision_client, image):
    img = vision.Image(content=image_to_bytes(image))
    resp = vision_client.document_text_detection(image=img)
    text = ""
    if getattr(resp, "full_text_annotation", None):
        text = getattr(resp.full_text_annotation, "text", "") or ""
    return text.strip()

def heuristic_parse_from_text(text):
    text = (text or "").strip()
    if not text:
        return {"bills": [{"shop_name": "Unknown Shop", "bill_date": None, "gst_number": None, "items": [], "total": "0"}]}
    
    lines = [x.strip() for x in text.splitlines() if x.strip()]
    shop_name = lines[0][:80] if lines else "Unknown Shop"
    
    return {
        "bills": [{
            "shop_name": shop_name,
            "bill_date": datetime.now().strftime("%Y-%m-%d"),
            "gst_number": "N/A",
            "items": [],
            "total": "0"
        }]
    }

def try_gemini(model, image):
    try:
        img_payload = {
            "data": image_to_bytes(image),
            "mime_type": "image/jpeg"
        }

        resp = model.generate_content(
            [build_schema_prompt(), img_payload],
            generation_config={
                "temperature": 0,
                "response_mime_type": "application/json"
            },
        )

        data = parse_json_from_response(getattr(resp, "text", ""))
        st.session_state.gemini_available = True

        if not isinstance(data, dict): return {"bills": []}, None
        bills = data.get("bills", [])
        if not isinstance(bills, list): bills = []

        return {"bills": bills}, None
    except Exception as e:
        if is_gemini_quota_error(e):
            st.session_state.gemini_available = False
            st.session_state.last_gemini_error_time = datetime.now()
        return {"bills": []}, e

def try_perplexity(client, image):
    b64 = base64.b64encode(image_to_bytes(image)).decode("utf-8")
    resp = client.chat.completions.create(
        model=secret_or_default("PERPLEXITY_MODEL", "sonar-pro"),
        messages=[
            {"role": "system", "content": build_schema_prompt()},
            {"role": "user", "content": [
                {"type": "text", "text": "Extract invoices from this image structured in the requested array structure."},
                {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64}"}},
            ]},
        ],
        temperature=0.0,
    )
    return parse_json_from_response(resp.choices[0].message.content)

def try_openai(client, image):
    b64 = base64.b64encode(image_to_bytes(image)).decode("utf-8")
    resp = client.chat.completions.create(
        model=secret_or_default("OPENAI_MODEL", "gpt-4o-mini"),
        messages=[
            {"role": "system", "content": build_schema_prompt()},
            {"role": "user", "content": [
                {"type": "text", "text": "Extract billing arrays from this snapshot."},
                {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64}"}},
            ]},
        ],
        response_format={"type": "json_object"},
    )
    return parse_json_from_response(resp.choices[0].message.content)

def analyze_with_auto_fallback(model_bundle, image, forced=None):
    image = preprocess_for_ocr(image)
    order = ["Gemini", "Google Vision OCR", "Perplexity", "OpenAI"]
    if forced in order:
        order = [forced] + [x for x in order if x != forced]

    for provider in order:
        try:
            if provider == "Gemini" and model_bundle.get("gemini") and can_try_gemini():
                data, _ = try_gemini(model_bundle["gemini"], image)
                if data and "bills" in data: return data
            elif provider == "Google Vision OCR" and model_bundle.get("vision_client"):
                text = extract_vision_text(model_bundle["vision_client"], image)
                if text: return heuristic_parse_from_text(text)
            elif provider == "Perplexity" and model_bundle.get("perplexity"):
                res = try_perplexity(model_bundle["perplexity"], image)
                if res and "bills" in res: return res
            elif provider == "OpenAI" and model_bundle.get("openai"):
                res = try_openai(model_bundle["openai"], image)
                if res and "bills" in res: return res
        except Exception:
            continue
    return {"bills": []}

def insert_bill(shop, date, gst, total, calc_total, status):
    with sqlite3.connect(DB_PATH, timeout=30) as conn:
        conn.execute(
            """
            INSERT OR IGNORE INTO bills (shop_name, bill_date, gst_number, total, calculated_total, status, timestamp)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (shop, date, gst, float(total), float(calc_total), status, datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        )
        conn.commit()

def sanitize_sheet_name(name, index=1, fallback="Sheet"):
    name = re.sub(r"[\[\]\*\?\/\\:]", "_", str(name)).strip()
    name = re.sub(r"\s+", " ", name)
    return f"B{index}_{name[:20]}" if name else f"{fallback}_{index}"

def build_excel_export(results):
    buffer = BytesIO()
    wb = openpyxl.Workbook()
    
    ws1 = wb.active
    ws1.title = "Summary Dashboard"
    ws1.views.sheetView[0].showGridLines = True

    navy_dark = "1F4E78"
    navy_zebra = "F2F4F8"
    white = "FFFFFF"
    gray_border = "D9D9D9"

    font_title = Font(name="Calibri", size=16, bold=True, color="1F4E78")
    font_section = Font(name="Calibri", size=12, bold=True, color="1F4E78")
    font_header = Font(name="Calibri", size=11, bold=True, color=white)
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)

    fill_header = PatternFill(start_color=navy_dark, end_color=navy_dark, fill_type="solid")
    fill_zebra = PatternFill(start_color=navy_zebra, end_color=navy_zebra, fill_type="solid")

    thin_side = Side(border_style="thin", color=gray_border)
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_total = Border(top=Side(border_style="thin", color="000000"), bottom=Side(border_style="double", color="000000"))

    ws1["A1"] = "AI OCR PROCESSOR - INVOICE BATCH SUMMARY"
    ws1["A1"].font = font_title
    ws1["A2"] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws1["A2"].font = Font(name="Calibri", size=11, italic=True)

    ws1["A4"] = "1. Document-wise Statement Breakdown"
    ws1["A4"].font = font_section

    headers_bill = ["Sr No.", "Source File / Page", "Shop / Vendor Name", "Date", "Original Total", "Calculated Total", "Status"]
    for col_num, h in enumerate(headers_bill, 1):
        cell = ws1.cell(row=5, column=col_num, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws2 = wb.create_sheet(title="Detailed Transactions")
    ws2.views.sheetView[0].showGridLines = True
    headers_det = ["Source File", "Shop Name", "Date", "Item Name / Particulars", "Qty", "Rate", "Extracted Amount"]
    for col_num, h in enumerate(headers_det, 1):
        cell = ws2.cell(row=1, column=col_num, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    det_idx = 2
    row_idx = 6
    rendered_fingerprints = set()

    for idx, item in enumerate(results):
        source = item.get("source", "Unknown")
        page = item.get("page", 1)
        d = item.get("data") or {}

        shop = str(d.get("shop_name") or "Unknown Shop").strip()
        date_str = str(d.get("bill_date") or datetime.now().strftime("%Y-%m-%d")).strip()
        orig_total = safe_float(d.get("total"))

        items = normalize_items(d.get("items"))
        calc_total = 0.0
        
        for it in items:
            amt = safe_float(it.get("amount")) or (safe_float(it.get("qty")) * safe_float(it.get("rate")))
            calc_total += amt

        if orig_total == 0.0 and calc_total > 0:
            orig_total = calc_total

        fingerprint = f"{shop}_{date_str}_{orig_total:.2f}"
        if fingerprint in rendered_fingerprints: continue
        rendered_fingerprints.add(fingerprint)

        for it in items:
            amt = safe_float(it.get("amount")) or (safe_float(it.get("qty")) * safe_float(it.get("rate")))
            ws2.cell(row=det_idx, column=1, value=source)
            ws2.cell(row=det_idx, column=2, value=shop)
            ws2.cell(row=det_idx, column=3, value=date_str)
            ws2.cell(row=det_idx, column=4, value=it.get("name", "Unknown"))
            ws2.cell(row=det_idx, column=5, value=safe_float(it.get("qty")))
            ws2.cell(row=det_idx, column=6, value=safe_float(it.get("rate")))
            ws2.cell(row=det_idx, column=7, value=amt).number_format = "₹#,##0.00"
            for c in range(1, 8):
                cell = ws2.cell(row=det_idx, column=c)
                cell.font = font_regular
                cell.border = border_all
                if det_idx % 2 == 1: cell.fill = fill_zebra
            det_idx += 1

        status = "Needs Review" if orig_total <= 0 else ("Matched" if abs(calc_total - orig_total) < 1 else "Mismatch")

        ws1.cell(row=row_idx, column=1, value=row_idx - 5).alignment = Alignment(horizontal="center")
        ws1.cell(row=row_idx, column=2, value=f"{source} (P.{page})")
        ws1.cell(row=row_idx, column=3, value=shop)
        ws1.cell(row=row_idx, column=4, value=date_str).alignment = Alignment(horizontal="center")
        
        c_orig = ws1.cell(row=row_idx, column=5, value=orig_total)
        c_orig.number_format = "₹#,##0.00"
        
        c_calc = ws1.cell(row=row_idx, column=6, value=calc_total)
        c_calc.number_format = "₹#,##0.00"
        
        ws1.cell(row=row_idx, column=7, value=status).alignment = Alignment(horizontal="center")

        for c in range(1, 8):
            cell = ws1.cell(row=row_idx, column=c)
            cell.font = font_regular
            cell.border = border_all
            if row_idx % 2 == 1: cell.fill = fill_zebra

        clean_shop_title = sanitize_sheet_name(shop, row_idx-5)
        ws_bill = wb.create_sheet(title=clean_shop_title)
        ws_bill.views.sheetView[0].showGridLines = True
        
        ws_bill.cell(row=1, column=1, value=f"Vendor: {shop}").font = font_bold
        ws_bill.cell(row=2, column=1, value=f"Date: {date_str}").font = font_regular
        
        headers_b = ["Item Name / Particulars", "Qty", "Rate", "Amount"]
        for col_v, hv in enumerate(headers_b, 1):
            h_cell = ws_bill.cell(row=4, column=col_v, value=hv)
            h_cell.font = font_header
            h_cell.fill = fill_header
            h_cell.alignment = Alignment(horizontal="center")

        sub_idx = 5
        for it in items:
            ws_bill.cell(row=sub_idx, column=1, value=it.get("name"))
            ws_bill.cell(row=sub_idx, column=2, value=safe_float(it.get("qty")))
            ws_bill.cell(row=sub_idx, column=3, value=safe_float(it.get("rate"))).number_format = "₹#,##0.00"
            amt_calc = safe_float(it.get("amount")) or (safe_float(it.get("qty")) * safe_float(it.get("rate")))
            ws_bill.cell(row=sub_idx, column=4, value=amt_calc).number_format = "₹#,##0.00"
            for col_v in range(1, 5):
                cell_v = ws_bill.cell(row=sub_idx, column=col_v)
                cell_v.font = font_regular
                cell_v.border = border_all
            sub_idx += 1
            
        ws_bill.cell(row=sub_idx, column=1, value="Total Summary").font = font_bold
        tot_cell = ws_bill.cell(row=sub_idx, column=4, value=calc_total)
        tot_cell.font = font_bold
        tot_cell.number_format = "₹#,##0.00"
        tot_cell.border = border_total

        for col_v in range(1, 5):
            ws_bill.column_dimensions[get_column_letter(col_v)].width = 22
        row_idx += 1

    if row_idx > 6:
        ws1.cell(row=row_idx, column=3, value="Grand Total Summary").font = font_bold
        cell_sum_e = ws1.cell(row=row_idx, column=5, value=f"=SUM(E6:E{row_idx-1})")
        cell_sum_e.font = font_bold
        cell_sum_e.number_format = "₹#,##0.00"
        cell_sum_e.border = border_total
        
        cell_sum_f = ws1.cell(row=row_idx, column=6, value=f"=SUM(F6:F{row_idx-1})")
        cell_sum_f.font = font_bold
        cell_sum_f.number_format = "₹#,##0.00"
        cell_sum_f.border = border_total

    for col in range(1, 8):
        ws1.column_dimensions[get_column_letter(col)].width = 20
        ws2.column_dimensions[get_column_letter(col)].width = 20

    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def build_batch_summary(results):
    rows = []
    seen_fingerprints = set()
    
    for item in results:
        d = item.get("data")
        if d:
            shop = str(d.get("shop_name") or "Unknown Shop").strip()
            bill_date = str(d.get("bill_date") or datetime.now().strftime("%Y-%m-%d")).strip()
            
            items = normalize_items(d.get("items"))
            calc_total = 0.0
            for it in items:
                calc_total += safe_float(it.get("amount")) or (safe_float(it.get("qty")) * safe_float(it.get("rate")))
                
            total = safe_float(d.get("total"))
            if total == 0.0 and calc_total > 0:
                total = calc_total

            fingerprint = f"{shop}_{bill_date}_{total:.2f}"
            if fingerprint in seen_fingerprints: continue
            seen_fingerprints.add(fingerprint)
            
            status = "Needs Review" if total <= 0 else ("Matched" if abs(calc_total - total) < 1 else "Mismatch")
            
            try: insert_bill(shop, bill_date, d.get("gst_number"), total, calc_total, status)
            except Exception: pass

            rows.append({
                "page": item.get("page"),
                "source": item.get("source"),
                "shop_name": shop,
                "bill_date": bill_date,
                "gst_number": d.get("gst_number") or "N/A",
                "bill_total": total,
                "calculated_total": calc_total,
                "difference": abs(calc_total - total),
                "status": status,
                "items_raw": items
            })
    return pd.DataFrame(rows)

def render_theme_toggle(location="main"):
    mode = st.radio("Theme", ["light", "dark"], horizontal=True, index=0 if st.session_state.get("theme_mode", "light") == "light" else 1, key=f"theme_radio_{location}")
    if mode != st.session_state.get("theme_mode", "light"):
        st.session_state["theme_mode"] = mode
        st.rerun()

def make_share_text(df=None):
    total = len(df) if df is not None and not df.empty else 0
    matched = int((df["status"] == "Matched").sum()) if total else 0
    mismatch = int((df["status"] == "Mismatch").sum()) if total else 0
    review = int((df["status"] == "Needs Review").sum()) if total else 0
    return f"{APP_TITLE}\nFiles Extracted: {total}\nMatched: {matched}\nMismatch: {mismatch}\nReview Required: {review}"

def share_whatsapp(text): return "https://wa.me/?text=" + urllib.parse.quote(text)
def share_telegram(text): return "https://t.me/share/url?url=&text=" + urllib.parse.quote(text)
def share_email(text, subject="AI Multi-Bill Summary Report"): return "mailto:?subject=" + urllib.parse.quote(subject) + "&body=" + urllib.parse.quote(text)

def render_upload_module():
    st.markdown("""
        <div class="deep-csc-header">
            <div class="branding-text">
                <h1>🧾 AI Intelligent Multi-Bill OCR Processor</h1>
                <p>One-Page Multi-Bill Segregation & Verified Transaction Ledger Extraction Engine.</p>
            </div>
            <div class="csc-meta-badge">📍 <b>Deep CSC</b><br>👤 Owner: Deepak | ID: 256423250015</div>
            <div class="branding-badge">Ultra Build</div>
        </div>
    """, unsafe_allow_html=True)

    tabs = st.tabs(["📷 Scan Pipeline", "🕘 History Logs", "⚙️ Preferences"])

    with tabs[0]:
        providers = ["Gemini", "Google Vision OCR", "Perplexity", "OpenAI"]
        st.session_state.selected_provider = st.selectbox("Select OCR Provider Architecture", providers, index=providers.index("Gemini"))

        uploaded_files = st.file_uploader(
            "Upload Single/Multi-Bill Invoices (Images or PDFs)",
            type=["jpg", "jpeg", "png", "pdf"],
            accept_multiple_files=True
        )

        col_a, col_b = st.columns(2)
        with col_a: process_now = st.button("Execute Batch Analysis Pipeline", use_container_width=True)
        with col_b: clear_state = st.button("Flush Batch Cache", use_container_width=True)

        if clear_state:
            st.session_state.processed_files = set()
            st.session_state.current_results = []
            st.rerun()

        model_bundle = {
            "vision_client": setup_google_vision(),
            "gemini": setup_gemini(),
            "perplexity": setup_perplexity(),
            "openai": setup_openai(),
        }

        if process_now and uploaded_files:
            new_results = []
            progress_bar = st.progress(0)
            status_text = st.empty()

            total_items = len(uploaded_files)
            for idx, f in enumerate(uploaded_files):
                status_text.markdown(f"**Processing File:** `{f.name}` ({idx+1}/{total_items})")
                f_bytes = f.read()
                
                if f.name.lower().endswith(".pdf"):
                    try:
                        images = convert_pdf_to_images(f_bytes)
                        for p_idx, img in enumerate(images, 1):
                            res = analyze_with_auto_fallback(model_bundle, img, forced=st.session_state.selected_provider)
                            for single_bill in res.get("bills", []):
                                new_results.append({"source": f.name, "page": p_idx, "data": single_bill})
                    except Exception as e:
                        st.error(f"Failed to decode PDF `{f.name}`: {str(e)}")
                else:
                    try:
                        img = Image.open(BytesIO(f_bytes)).convert("RGB")
                        res = analyze_with_auto_fallback(model_bundle, img, forced=st.session_state.selected_provider)
                        for single_bill in res.get("bills", []):
                            new_results.append({"source": f.name, "page": 1, "data": single_bill})
                    except Exception as e:
                        st.error(f"Failed processing image `{f.name}`: {str(e)}")

                progress_bar.progress(int((idx + 1) / total_items * 100))

            st.session_state.current_results = new_results
            status_text.success("🟢 Complete batch sequence parsed successfully.")

        if st.session_state.current_results:
            df = build_batch_summary(st.session_state.current_results)
            render_metrics(df)

            st.markdown("### 📊 Live Extracted Invoice Statement Summary")
            
            def style_status(val):
                if val == "Matched": return "background-color: #d1fae5; color: #065f46; font-weight: bold;"
                if val == "Mismatch": return "background-color: #fee2e2; color: #991b1b; font-weight: bold;"
                return "background-color: #fef3c7; color: #92400e;"

            if not df.empty:
                display_df = df[["source", "page", "shop_name", "bill_date", "gst_number", "bill_total", "calculated_total", "status"]]
                st.dataframe(display_df.style.map(style_status, subset=["status"]), use_container_width=True)

                col_down1, col_down2 = st.columns(2)
                with col_down1:
                    excel_bytes = build_excel_export(st.session_state.current_results)
                    st.download_button(
                        label="📥 Download Formatted Excel Ledger (.xlsx)",
                        data=excel_bytes,
                        file_name=f"Deep_CSC_Batch_Ledger_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                with col_down2:
                    report_txt = make_share_text(df)
                    st.markdown(f"""
                        <div style="display:flex; gap:10px; margin-top:2px;">
                            <a href="{share_whatsapp(report_txt)}" target="_blank" style="flex:1; text-align:center; background:#25D366; color:white; padding:10px; border-radius:12px; text-decoration:none; font-weight:bold;">WhatsApp Report</a>
                            <a href="{share_telegram(report_txt)}" target="_blank" style="flex:1; text-align:center; background:#0088cc; color:white; padding:10px; border-radius:12px; text-decoration:none; font-weight:bold;">Telegram Report</a>
                            <a href="{share_email(report_txt)}" target="_blank" style="flex:1; text-align:center; background:#ea4335; color:white; padding:10px; border-radius:12px; text-decoration:none; font-weight:bold;">Email Audit</a>
                        </div>
                    """, unsafe_allow_html=True)
            else:
                st.info("No explicit structured bills extracted from current data views.")

    with tabs[1]:
        st.markdown("### 🏢 Historical Synchronization Logs")
        with sqlite3.connect(DB_PATH) as conn:
            hist_df = pd.read_sql_query("SELECT * FROM bills ORDER BY timestamp DESC LIMIT 200", conn)
        if not hist_df.empty:
            st.dataframe(hist_df, use_container_width=True)
        else:
            st.write("Database index logs are currently empty.")

    with tabs[2]:
        st.markdown("### ⚙️ Master Environment Parameters")
        render_theme_toggle("tab")

def render_sidebar():
    with st.sidebar:
        st.markdown(f"### {APP_TITLE}")
        st.markdown("---")
        st.markdown(f"**Session Status:** `Authenticated`")
        if st.button("Secure Logout"):
            terminate_session()

def main():
    setup_page()
    init_db()
    init_auth()
    init_runtime_state()
    apply_theme_css()

    if not st.session_state.logged_in:
        do_login()
    else:
        apply_css()
        render_sidebar()
        render_upload_module()

if __name__ == "__main__":
    main()
